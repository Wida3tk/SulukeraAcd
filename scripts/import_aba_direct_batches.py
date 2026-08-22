from __future__ import annotations

import csv
import hashlib
import json
import random
import re
import string
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INDEX_HTML = ROOT / "index.html"
WORKBOOK_PATH = Path(r"C:\Users\Maryam\Downloads\قاعدة بيانات طلاب منصة سلوكيرا.xlsx")
TODAY = date(2026, 8, 22).isoformat()

PROGRAM_NAME = "تحليل السلوك التطبيقي"
BATCH_MAP = {10: "Q1-26", 11: "Q2-26"}
PKG_MAP = {"QBA": "pkg_1777063589271", "QASP-S": "pkg_1777064163359"}
PASSWORD_CHARS = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"

BOARD_SHEET_NAME = "لوحة بيانات ABA مباشر"
CARD_ALIASES = {
    "رفيف عبد الرحمن السدراني": "رفيف عبد الرحمن السدراني",
    "دلال الجهني": "دلال غنايم الجهني",
    "وجدان الحارثي": "وجدان سعيد الحارثي",
    "ريم القفاري": "ريم محمد القفاري",
    "خديجة المؤمن": "خديجه عبدالله المومن",
}


def read_firebase_base_url() -> str:
    html = INDEX_HTML.read_text(encoding="utf-8")
    match = re.search(r'databaseURL:\s*"([^"]+)"', html)
    if not match:
        raise RuntimeError("databaseURL not found in index.html")
    return match.group(1).rstrip("/")


BASE_URL = read_firebase_base_url()


def get_json(path: str) -> Any:
    with urllib.request.urlopen(f"{BASE_URL}/{path}.json") as response:
        return json.load(response)


def patch_json(path: str, payload: dict[str, Any]) -> None:
    req = urllib.request.Request(
        f"{BASE_URL}/{path}.json",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="PATCH",
    )
    with urllib.request.urlopen(req):
        return


def canon_name(value: str) -> str:
    text = str(value or "").strip().lower()
    table = str.maketrans(
        {
            "أ": "ا",
            "إ": "ا",
            "آ": "ا",
            "ٱ": "ا",
            "ة": "ه",
            "ى": "ي",
            "ؤ": "و",
            "ئ": "ي",
        }
    )
    text = text.translate(table)
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").strip())


def normalize_email(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_phone(value: Any) -> str:
    return "".join(ch for ch in str(value or "").replace(".0", "") if ch.isdigit())


def split_name(full_name: str) -> tuple[str, str, str]:
    parts = [part for part in normalize_text(full_name).split(" ") if part]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def random_password(length: int = 10) -> str:
    rng = random.SystemRandom()
    return "".join(rng.choice(PASSWORD_CHARS) for _ in range(length))


def sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def password_fields(password: str) -> dict[str, str | None]:
    salt = random_password(16)
    return {"passwordHash": sha256(salt + password), "passwordSalt": salt, "password": None}


def make_empty_grade() -> dict[str, Any]:
    return {
        "w1_attend": 0,
        "w1_hw": 0,
        "w1_disc": 0,
        "w2_attend": 0,
        "w2_hw": 0,
        "w2_disc": 0,
        "w3_attend": 0,
        "w3_hw": 0,
        "w3_disc": 0,
        "w4_attend": 0,
        "w4_hw": 0,
        "w4_disc": 0,
        "exam": None,
        "examEntered": False,
        "updatedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }


@dataclass
class Candidate:
    batch_no: int
    batch_code: str
    name: str
    license: str
    email: str = ""
    phone: str = ""
    card_sheet: str = ""
    card_found: bool = False
    missing_details: bool = False
    duplicate_reason: str = ""


def board_candidates(workbook) -> list[Candidate]:
    sheet = workbook[BOARD_SHEET_NAME]
    candidates: list[Candidate] = []
    for row in range(13, 80):
        left_name = normalize_text(sheet.cell(row, 4).value)
        left_license = normalize_text(sheet.cell(row, 7).value).upper()
        if left_name:
            candidates.append(Candidate(batch_no=10, batch_code=BATCH_MAP[10], name=left_name, license=left_license))
        right_name = normalize_text(sheet.cell(row, 22).value)
        right_license = normalize_text(sheet.cell(row, 25).value).upper()
        if right_name:
            candidates.append(Candidate(batch_no=11, batch_code=BATCH_MAP[11], name=right_name, license=right_license))
    return candidates


def build_card_lookup(workbook) -> dict[str, dict[str, Any]]:
    cards: dict[str, dict[str, Any]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        student_name = normalize_text(sheet["C3"].value)
        email = normalize_email(sheet["K3"].value)
        phone = normalize_phone(sheet["I3"].value)
        license_name = normalize_text(sheet["N3"].value).upper()
        if not student_name or not (email or phone or license_name):
            continue
        cards[canon_name(student_name)] = {
            "sheet": sheet_name,
            "name": student_name,
            "email": email,
            "phone": phone,
            "license": license_name,
        }
    return cards


def attach_cards(candidates: list[Candidate], cards: dict[str, dict[str, Any]]) -> None:
    for candidate in candidates:
        card = cards.get(canon_name(candidate.name))
        if not card:
            alias = CARD_ALIASES.get(candidate.name)
            if alias:
                card = cards.get(canon_name(alias))
        if not card:
            for key, value in cards.items():
                tokens_a = set(canon_name(candidate.name).split())
                tokens_b = set(key.split())
                if len(tokens_a & tokens_b) >= max(2, min(len(tokens_a), len(tokens_b)) - 1):
                    card = value
                    break
        if card:
            candidate.card_found = True
            candidate.card_sheet = card["sheet"]
            candidate.email = card["email"]
            candidate.phone = card["phone"]
            if not candidate.license:
                candidate.license = card["license"]
        else:
            candidate.missing_details = True


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    students_node = get_json("students") or {}
    users_node = get_json("users") or {}
    enrollments_node = get_json("enrollments") or {}
    grades_node = get_json("grades") or {}
    subjects_node = get_json("subjects") or {}

    backup_dir = ROOT / "backups"
    backup_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"before-import-aba-direct-{stamp}.json"
    backup_path.write_text(
        json.dumps(
            {
                "students": students_node,
                "users": users_node,
                "enrollments": enrollments_node,
                "grades": grades_node,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    existing_students = list(students_node.values())
    existing_names = {canon_name(" ".join(filter(None, [s.get("firstName"), s.get("midName"), s.get("lastName")]))): s for s in existing_students}
    existing_emails = {normalize_email(s.get("email")): s for s in existing_students if normalize_email(s.get("email"))}
    existing_phones = {normalize_phone(s.get("phone")): s for s in existing_students if normalize_phone(s.get("phone"))}

    subjects_by_batch: dict[str, list[tuple[str, dict[str, Any]]]] = {}
    for subject_key, subject in subjects_node.items():
        batch = subject.get("batch") or ""
        subjects_by_batch.setdefault(batch, []).append((subject_key, subject))

    candidates = board_candidates(workbook)
    attach_cards(candidates, build_card_lookup(workbook))

    max_id = 0
    for student in existing_students:
        match = re.match(r"SUL-(\d+)$", str(student.get("id", "")))
        if match:
            max_id = max(max_id, int(match.group(1)))

    new_students_payload: dict[str, Any] = {}
    new_users_payload: dict[str, Any] = {}
    new_enrollments_payload: dict[str, Any] = {}
    new_grades_payload: dict[str, Any] = {}
    password_rows: list[list[str]] = [["اسم الطالب", "الرقم الطلابي", "كلمة المرور المؤقتة", "الدفعة", "المسار", "حالة البيانات"]]
    report_rows: list[dict[str, Any]] = []

    added_count = 0
    skipped_duplicates = 0
    missing_count = 0

    for candidate in candidates:
        canonical = canon_name(candidate.name)
        email = normalize_email(candidate.email)
        phone = normalize_phone(candidate.phone)
        duplicate_reason = ""
        if canonical in existing_names:
            duplicate_reason = "name"
        elif email and email in existing_emails:
            duplicate_reason = "email"
        elif phone and phone in existing_phones:
            duplicate_reason = "phone"
        if duplicate_reason:
            skipped_duplicates += 1
            report_rows.append(
                {
                    "name": candidate.name,
                    "batch": candidate.batch_code,
                    "track": candidate.license or "",
                    "status": "skipped_duplicate",
                    "reason": duplicate_reason,
                }
            )
            continue

        max_id += 1
        student_id = f"SUL-{max_id:03d}"
        first_name, mid_name, last_name = split_name(candidate.name)
        license_name = candidate.license if candidate.license in PKG_MAP else "QASP-S"
        pkg = PKG_MAP[license_name]
        notes = [
            f"تمت الإضافة من لوحة ABA مباشر بتاريخ {TODAY}.",
            "البيانات المالية بانتظار اعتماد موظف المالية قبل الاعتماد النهائي للطالب/الطالبة.",
        ]
        if candidate.missing_details:
            notes.append("هذا السجل يحتاج استكمالًا إداريًا للبيانات الناقصة من قبل الإدارة.")
            missing_count += 1
        if candidate.card_sheet:
            notes.append(f"بطاقة المصدر: {candidate.card_sheet}")

        student_key = f"std_{student_id.replace('-', '_')}"
        student_payload = {
            "firstName": first_name,
            "midName": mid_name,
            "lastName": last_name,
            "phone": phone,
            "email": email,
            "regDate": TODAY,
            "program": PROGRAM_NAME,
            "track": license_name,
            "batch": candidate.batch_code,
            "id": student_id,
            "pkg": pkg,
            "planType": license_name,
            "financeMode": "local_manual",
            "totalAmount": 0,
            "notes": " ".join(notes),
            "operCustomerId": "",
            "operOrderId": "",
            "installments": {
                "inst_0": {
                    "label": "قسط 1",
                    "amount": 0,
                    "due": "",
                    "status": "pending",
                }
            },
        }
        new_students_payload[student_key] = student_payload

        temp_password = random_password(10)
        user_key = f"user_{student_id.replace('-', '_')}"
        new_users_payload[user_key] = {
            "name": candidate.name,
            "username": student_id,
            "role": "student",
            **password_fields(temp_password),
        }
        password_rows.append(
            [
                candidate.name,
                student_id,
                temp_password,
                candidate.batch_code,
                license_name,
                "بيانات ناقصة" if candidate.missing_details else "مكتمل",
            ]
        )

        for subject_key, subject in subjects_by_batch.get(candidate.batch_code, []):
            eligible = subject.get("eligiblePlans") or []
            if eligible and license_name not in eligible:
                continue
            enrollment_key = f"enr_{student_key}_{subject_key}".replace(".", "_")
            new_enrollments_payload[enrollment_key] = {"studentKey": student_key, "subjectKey": subject_key}
            grade_key = f"grade_{student_key}_{subject_key}".replace(".", "_")
            grade_payload = make_empty_grade()
            grade_payload["studentKey"] = student_key
            grade_payload["subjectKey"] = subject_key
            new_grades_payload[grade_key] = grade_payload

        existing_names[canonical] = student_payload
        if email:
            existing_emails[email] = student_payload
        if phone:
            existing_phones[phone] = student_payload
        added_count += 1
        report_rows.append(
            {
                "name": candidate.name,
                "batch": candidate.batch_code,
                "track": license_name,
                "status": "added_missing_details" if candidate.missing_details else "added",
                "studentId": student_id,
                "cardSheet": candidate.card_sheet,
            }
        )

    if new_students_payload:
        patch_json("students", new_students_payload)
        patch_json("users", new_users_payload)
        patch_json("enrollments", new_enrollments_payload)
        patch_json("grades", new_grades_payload)

    report_dir = ROOT / "reports"
    report_dir.mkdir(exist_ok=True)
    passwords_path = report_dir / f"sulukera-import-passwords-{stamp}.csv"
    with passwords_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(password_rows)

    report_path = report_dir / f"sulukera-import-report-{stamp}.json"
    report_path.write_text(
        json.dumps(
            {
                "addedCount": added_count,
                "missingDetailsCount": missing_count,
                "skippedDuplicates": skipped_duplicates,
                "backup": str(backup_path),
                "passwordsCsv": str(passwords_path),
                "rows": report_rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(json.dumps({"addedCount": added_count, "missingDetailsCount": missing_count, "skippedDuplicates": skipped_duplicates, "backup": str(backup_path), "passwordsCsv": str(passwords_path), "report": str(report_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
